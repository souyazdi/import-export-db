# -*- coding: utf-8 -*-
"""
Created on Mon Dec 16 18:06:21 2019

@author: yazdsous
"""
import pyodbc
import pandas as pd
import xlrd
from openpyxl import load_workbook

conn0 = pyodbc.connect('Driver={SQL Server};'
                      'Server=Hosaka\Sqlp2;'
                      'Database=Eforms;'
                      'Trusted_Connection=yes;')


Z = "SELECT [FormId]\
      ,[Name]\
      ,[FilingId]\
      ,[AddedOn]\
  FROM [Eforms].[dbo].[Form]\
  WHERE FilingID IS NOT NULL AND [Name] = N's15ab_ShrtTrmNtrlGs_ImprtExprt'\
  ORDER BY FormId DESC"


df0 = pd.read_sql(Z,conn0)

df1 = df0.sort_values(by=['FormId'], ascending = True)     

df2 = df1[-50:]

###############################################################################

file = "oas_forms.xlsx"

writer = pd.ExcelWriter(file, engine='openpyxl')

try:
        # try to open an existing workbook
        writer.book = load_workbook(file)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        startrow = writer.book['Sheet1'].max_row


        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index('Sheet1')
        
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
        
        # write out the new sheet
        df2.to_excel(writer,sheet_name='Sheet1', startrow=startrow, index = False,header= True)

# save the workbook
        writer.save()

except FileNotFoundError:
        # file does not exist yet, we will create it
        pass



import xlrd
filename ='oas_forms.xls'
book = xlrd.open_workbook(filename)
sheet = book.sheet_by_name(mysheet) 
df = pd.read_excel("oas_forms.xlsx")  

  

